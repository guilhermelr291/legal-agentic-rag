"""XLSX metadata extractor using openpyxl."""

from dataclasses import dataclass, field

from openpyxl import load_workbook


@dataclass
class XLSXMetadata:
    """Metadata extracted from an XLSX file.

    Attributes:
        sheet_names: List of sheet names
        sheets: Detailed info for each sheet (columns, row count, etc.)
        total_sheets: Total number of sheets
        file_path: Path to the source file
        error: Error message if extraction failed
        success: Whether extraction succeeded
    """

    sheet_names: list[str] = field(default_factory=list)
    sheets: list[dict] = field(default_factory=list)
    total_sheets: int = 0
    file_path: str = ""
    error: str | None = None
    success: bool = True


class XLSXMetadataExtractor:
    """Extract metadata from XLSX files (not full text)."""

    def extract(self, file_path: str) -> XLSXMetadata:
        """Extract metadata from an XLSX file.

        Args:
            file_path: Path to the XLSX file

        Returns:
            XLSXMetadata with sheet names, columns, and row counts
        """
        try:
            # Load workbook in read-only mode for efficiency
            wb = load_workbook(file_path, read_only=True, data_only=True)

            sheet_names = wb.sheetnames
            sheets_info: list[dict] = []

            for sheet_name in sheet_names:
                sheet = wb[sheet_name]

                # Get column headers from first row
                columns: list[str] = []
                row_count = 0
                has_headers = False

                # Iterate rows to count and extract headers
                for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                    if row_idx == 1:
                        # First row - treat as headers
                        columns = [
                            str(cell.value) if cell.value else f"Column_{i + 1}"
                            for i, cell in enumerate(row)
                        ]
                        has_headers = any(cell.value is not None for cell in row)
                    else:
                        # Count non-empty rows
                        if any(cell.value is not None for cell in row):
                            row_count += 1

                sheet_info: dict = {
                    "name": sheet_name,
                    "columns": columns,
                    "column_count": len(columns),
                    "row_count": row_count,
                    "has_headers": has_headers,
                }

                # Add dimension info if available
                if sheet.dimensions:
                    sheet_info["dimensions"] = sheet.dimensions

                sheets_info.append(sheet_info)

            wb.close()

            return XLSXMetadata(
                sheet_names=sheet_names,
                sheets=sheets_info,
                total_sheets=len(sheet_names),
                file_path=file_path,
                success=True,
            )

        except FileNotFoundError:
            return XLSXMetadata(
                error=f"File not found: {file_path}",
                success=False,
            )
        except PermissionError:
            return XLSXMetadata(
                error=f"Permission denied: {file_path}",
                success=False,
            )
        except Exception as e:
            return XLSXMetadata(
                error=f"XLSX extraction failed: {e!s}",
                success=False,
            )

    def extract_raw(self, file_path: str) -> XLSXMetadata:
        """Alias for extract() - for interface compatibility.

        Args:
            file_path: Path to the XLSX file

        Returns:
            XLSXMetadata with sheet names, columns, and row counts
        """
        return self.extract(file_path)
